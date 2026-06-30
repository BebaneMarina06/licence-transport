from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.application import Application, ApplicationStatus, DeliveryFormat
from app.schemas.dashboard import DashboardStats
from app.schemas.revenue import RevenueEntryResponse, RevenueSummaryResponse

HEADER_FILL = PatternFill("solid", fgColor="0055A4")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="0055A4")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="334155")
BODY_FONT = Font(name="Calibri", size=11)
META_FONT = Font(name="Calibri", size=10, color="64748B")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
CURRENCY_FMT = '#,##0 "XAF"'
DATE_FMT = "DD/MM/YYYY HH:MM"
INTEGER_FMT = "#,##0"
DECIMAL_FMT = "0.0"

APPLICATION_STATUS_LABELS: dict[str, str] = {
    ApplicationStatus.DRAFT.value: "Brouillon",
    ApplicationStatus.SUBMITTED.value: "Soumis",
    ApplicationStatus.UNDER_REVIEW.value: "En instruction",
    ApplicationStatus.COMPLEMENT_REQUESTED.value: "Complément demandé",
    ApplicationStatus.APPROVED.value: "Approuvé",
    ApplicationStatus.AWAITING_PAYMENT.value: "En attente de paiement",
    ApplicationStatus.PAID.value: "Payé",
    ApplicationStatus.DELIVERED.value: "Délivré",
    ApplicationStatus.REJECTED.value: "Rejeté",
    ApplicationStatus.CANCELLED.value: "Annulé",
}

DELIVERY_FORMAT_LABELS: dict[str, str] = {
    DeliveryFormat.DIGITAL.value: "Numérique",
    DeliveryFormat.PHYSICAL.value: "Physique",
}

REVENUE_STATE_LABELS: dict[str, str] = {
    "confirmed": "Confirmé",
    "pending_validation": "En attente de validation",
    "other": "Autre",
}


def status_label(status: ApplicationStatus | str, *, amount_paid: float | None = None) -> str:
    key = status.value if isinstance(status, ApplicationStatus) else status
    if key == ApplicationStatus.AWAITING_PAYMENT.value and amount_paid is not None:
        return "En attente de confirmation"
    return APPLICATION_STATUS_LABELS.get(key, key)


def delivery_format_label(value: str | None) -> str:
    if not value:
        return ""
    return DELIVERY_FORMAT_LABELS.get(value, value)


def revenue_state_label(value: str) -> str:
    return REVENUE_STATE_LABELS.get(value, value)


def _excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is not None:
        return value.astimezone(UTC).replace(tzinfo=None)
    return value


def _style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body_cell(cell, *, alt: bool = False) -> None:
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    if alt:
        cell.fill = ALT_FILL


def _auto_width(ws, min_width: int = 12, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max((len(str(cell.value or "")) for cell in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def _write_report_header(ws, title: str, ncol: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    generated = datetime.now(UTC).strftime("%d/%m/%Y à %H:%M UTC")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    meta_cell = ws.cell(row=2, column=1, value=f"Généré le {generated}")
    meta_cell.font = META_FONT
    return 4


def _write_section_title(ws, row: int, title: str, ncol: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    return row + 1


def _write_key_value_table(
    ws,
    start_row: int,
    rows: list[tuple[str, Any]],
    *,
    alt: bool = False,
    currency_labels: set[str] | None = None,
    decimal_labels: set[str] | None = None,
) -> int:
    currency_labels = currency_labels or set()
    decimal_labels = decimal_labels or set()
    row = start_row
    for idx, (label, value) in enumerate(rows):
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value)
        is_alt = alt or idx % 2 == 1
        _style_body_cell(label_cell, alt=is_alt)
        _style_body_cell(value_cell, alt=is_alt)
        if label in currency_labels and isinstance(value, (int, float)):
            value_cell.number_format = CURRENCY_FMT
        elif label in decimal_labels and isinstance(value, (int, float)):
            value_cell.number_format = DECIMAL_FMT
        elif isinstance(value, int):
            value_cell.number_format = INTEGER_FMT
        row += 1
    return row


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_stats_workbook(stats: DashboardStats) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False

    row = _write_report_header(ws, "Rapport statistiques — Licences de transport DGTT", 2)

    row = _write_section_title(ws, row, "Indicateurs généraux", 2)
    row = _write_key_value_table(
        ws,
        row,
        [
            ("Total dossiers", stats.total_applications),
            ("En instruction", stats.pending_review),
            ("En attente de paiement", stats.awaiting_payment),
            ("Délivrés", stats.delivered),
            ("Rejetés", stats.rejected),
            ("Citoyens inscrits", stats.total_citizens),
            ("Recettes totales", stats.total_revenue),
            ("Recettes du mois", stats.revenue_this_month),
            (
                "Délai moyen de traitement (jours)",
                stats.avg_processing_days if stats.avg_processing_days is not None else "—",
            ),
            ("Dossiers en retard", stats.overdue_count),
        ],
        currency_labels={"Recettes totales", "Recettes du mois"},
        decimal_labels={"Délai moyen de traitement (jours)"},
    )

    row += 1
    row = _write_section_title(ws, row, "Répartition par statut", 2)
    status_rows = [
        (status_label(name), count)
        for name, count in sorted(stats.applications_by_status.items(), key=lambda item: item[1], reverse=True)
    ]
    row = _write_key_value_table(ws, row, status_rows, alt=True)

    row += 1
    row = _write_section_title(ws, row, "Répartition par type de licence", 2)
    type_rows = [
        (name, count)
        for name, count in sorted(stats.applications_by_license_type.items(), key=lambda item: item[1], reverse=True)
    ]
    _write_key_value_table(ws, row, type_rows, alt=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    return _workbook_to_bytes(wb)


def build_applications_workbook(applications: list[Application]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dossiers"
    ws.sheet_view.showGridLines = False

    headers = [
        "Référence",
        "Statut",
        "Type de licence",
        "Demandeur",
        "Entreprise",
        "Immatriculation",
        "Montant payé (XAF)",
        "Soumis le",
        "Créé le",
    ]
    ncol = len(headers)
    _write_report_header(ws, "Liste des dossiers — Licences de transport DGTT", ncol)

    header_row = 4
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)
    _style_header_row(ws, header_row, ncol)
    ws.freeze_panes = "A5"

    for idx, app in enumerate(applications):
        row = header_row + 1 + idx
        alt = idx % 2 == 1
        values: list[Any] = [
            app.reference,
            status_label(app.status, amount_paid=float(app.amount_paid) if app.amount_paid is not None else None),
            app.license_type.name if app.license_type else "",
            app.applicant.full_name if app.applicant else "",
            app.company_name or "",
            app.vehicle_plate or "",
            float(app.amount_paid) if app.amount_paid is not None else None,
            _excel_datetime(app.submitted_at),
            _excel_datetime(app.created_at),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            _style_body_cell(cell, alt=alt)
            if col == 7 and isinstance(value, (int, float)):
                cell.number_format = CURRENCY_FMT
            elif col in (8, 9) and value is not None:
                cell.number_format = DATE_FMT

    _auto_width(ws)
    return _workbook_to_bytes(wb)


def build_revenue_workbook(
    summary: RevenueSummaryResponse,
    entries: list[RevenueEntryResponse],
) -> bytes:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Synthèse"
    ws_summary.sheet_view.showGridLines = False
    row = _write_report_header(ws_summary, "Synthèse des recettes — Licences de transport DGTT", 2)
    row = _write_section_title(ws_summary, row, "Indicateurs financiers", 2)
    row = _write_key_value_table(
        ws_summary,
        row,
        [
            ("Recettes confirmées", summary.total_confirmed),
            ("Recettes du mois", summary.confirmed_this_month),
            ("En attente de validation", summary.pending_validation_amount),
            ("Paiements confirmés (nombre)", summary.confirmed_count),
            ("Dossiers en attente de paiement", summary.awaiting_payment_count),
        ],
        currency_labels={
            "Recettes confirmées",
            "Recettes du mois",
            "En attente de validation",
        },
    )

    row += 1
    row = _write_section_title(ws_summary, row, "Recettes par type de licence", 3)
    for col, header in enumerate(["Type de licence", "Nombre", "Montant (XAF)"], start=1):
        ws_summary.cell(row=row, column=col, value=header)
    _style_header_row(ws_summary, row, 3)
    row += 1
    for idx, item in enumerate(summary.by_license_type):
        alt = idx % 2 == 1
        for col, value in enumerate([item.license_type_name, item.count, item.total_amount], start=1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            _style_body_cell(cell, alt=alt)
            if col == 3:
                cell.number_format = CURRENCY_FMT
            elif col == 2:
                cell.number_format = INTEGER_FMT
        row += 1

    row += 1
    row = _write_section_title(ws_summary, row, "Recettes par mois", 3)
    for col, header in enumerate(["Mois", "Nombre", "Montant (XAF)"], start=1):
        ws_summary.cell(row=row, column=col, value=header)
    _style_header_row(ws_summary, row, 3)
    row += 1
    for idx, item in enumerate(summary.by_month):
        alt = idx % 2 == 1
        for col, value in enumerate([item.month, item.count, item.total_amount], start=1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            _style_body_cell(cell, alt=alt)
            if col == 3:
                cell.number_format = CURRENCY_FMT
            elif col == 2:
                cell.number_format = INTEGER_FMT
        row += 1

    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 20

    ws_detail = wb.create_sheet("Détail")
    ws_detail.sheet_view.showGridLines = False
    detail_headers = [
        "Référence",
        "Demandeur",
        "Type de licence",
        "Montant (XAF)",
        "Format de délivrance",
        "Statut du dossier",
        "État de la recette",
        "Date de paiement",
        "Référence de paiement",
    ]
    ncol = len(detail_headers)
    _write_report_header(ws_detail, "Détail des recettes", ncol)
    header_row = 4
    for col, header in enumerate(detail_headers, start=1):
        ws_detail.cell(row=header_row, column=col, value=header)
    _style_header_row(ws_detail, header_row, ncol)
    ws_detail.freeze_panes = "A5"

    for idx, entry in enumerate(entries):
        row = header_row + 1 + idx
        alt = idx % 2 == 1
        values: list[Any] = [
            entry.reference,
            entry.applicant_name,
            entry.license_type_name,
            entry.amount,
            delivery_format_label(entry.delivery_format),
            status_label(entry.status, amount_paid=entry.amount),
            revenue_state_label(entry.revenue_state),
            _excel_datetime(entry.paid_at),
            entry.payment_reference or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws_detail.cell(row=row, column=col, value=value)
            _style_body_cell(cell, alt=alt)
            if col == 4:
                cell.number_format = CURRENCY_FMT
            elif col == 8 and value is not None:
                cell.number_format = DATE_FMT

    _auto_width(ws_detail)
    return _workbook_to_bytes(wb)
